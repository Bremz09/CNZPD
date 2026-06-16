import argparse
import re
from pathlib import Path

import pandas as pd
import trueskill
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


REQUIRED_COLUMNS = [
    "Location",
    "Year",
    "Date",
    "Event",
    "Round",
    "Heat",
    "Rank",
    "race_id_name",
    "race_id",
    "Athlete",
    "Country",
    "Age",
]

OUTPUT_COLUMNS = [
    "Final_Rank",
    "Initial_Mu",
    "Initial_Sigma",
    "Initial_CSE",
    "Final_Mu",
    "Final_Sigma",
    "Final_CSE",
    "Strength_of_Field",
]

SCRIPT_DIR = Path(__file__).resolve().parent
CNZPD_DIR = SCRIPT_DIR.parent
KEIRIN_BETA = 5.2083
KEIRIN_TAU = 0.75


def normalize_round(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip().upper()
    return " ".join(text.split())


def parse_heat_number(value: object) -> int:
    if pd.isna(value):
        return 1

    text = str(value).strip()
    if not text:
        return 1

    try:
        number = int(float(text))
        return number if number > 0 else 1
    except ValueError:
        match = re.search(r"(\d+)", text)
        if match:
            return int(match.group(1))
    return 1


def darken_hex_color(hex_color: str, steps: int, per_step: float = 0.03) -> str:
    color = hex_color.strip().lstrip("#")
    if len(color) != 6:
        return hex_color

    ratio = max(0.0, 1.0 - max(0, steps) * per_step)
    r = int(color[0:2], 16)
    g = int(color[2:4], 16)
    b = int(color[4:6], 16)

    dr = int(round(r * ratio))
    dg = int(round(g * ratio))
    db = int(round(b * ratio))
    return f"{dr:02X}{dg:02X}{db:02X}"


def keirin_row_fill(round_value: object, heat_value: object) -> PatternFill | None:
    round_key = normalize_round(round_value)
    heat_num = parse_heat_number(heat_value)

    if round_key in {"F", "FINAL"}:
        color = darken_hex_color("D9E1F2", heat_num - 1)
    elif round_key in {"R1", "ROUND 1"}:
        color = darken_hex_color("E2EFDA", heat_num - 1)
    elif round_key in {"R1 REP", "R1 REPECHAGE", "ROUND 1 REP"}:
        color = darken_hex_color("E2EFDA", heat_num - 1)
    elif round_key in {"R2", "ROUND 2"}:
        color = darken_hex_color("FCE4D6", heat_num - 1)
    elif round_key in {"R2 REP", "R2 REPECHAGE", "ROUND 2 REP"}:
        color = darken_hex_color("FCE4D6", heat_num - 1)
    elif round_key in {"R3", "ROUND 3"}:
        color = darken_hex_color("FFF2CC", heat_num - 1)
    else:
        return None

    argb = f"FF{color}"
    return PatternFill(fill_type="solid", start_color=argb, end_color=argb)


def validate_columns(df: pd.DataFrame) -> None:
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(
            "Input sheet is missing required columns: " + ", ".join(missing)
        )


def rank_values_for_trueskill(ranks: pd.Series) -> list[int]:
    parsed = pd.to_numeric(ranks, errors="coerce")
    fallback = 1
    clean: list[int] = []

    for value in parsed:
        if pd.isna(value):
            clean.append(fallback)
        else:
            clean.append(int(value))
        fallback += 1

    return clean


def is_excluded_rank(rank: object) -> bool:
    if pd.isna(rank):
        return False
    text = str(rank).strip().upper()
    return text in {"DNF", "DNS", "REL"}


def _string_part(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return text


def _int_string_part(value: object) -> str:
    if pd.isna(value):
        return ""
    try:
        return str(int(float(value)))
    except Exception:  # noqa: BLE001
        return _string_part(value)


def _build_race_id_name_from_row(row: pd.Series) -> str:
    parts = [
        _string_part(row.get("Location", "")),
        _int_string_part(row.get("Year", "")),
        _string_part(row.get("Event", "")),
        _int_string_part(row.get("Round", "")),
        _int_string_part(row.get("Heat", "")),
    ]
    clean_parts = []
    for value in parts:
        text = str(value).strip()
        if text:
            clean_parts.append(text)
    return " ".join(clean_parts)


def fill_required_fields(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()

    generated_race_id_name = work.apply(_build_race_id_name_from_row, axis=1).astype("string")
    race_id_fallback = work["race_id"].astype("string").fillna("").str.strip()

    race_id_name_series = generated_race_id_name
    race_id_name_series = race_id_name_series.where(race_id_name_series != "", race_id_fallback)
    work["race_id_name"] = race_id_name_series

    return work


def compute_keirin_trueskill(df: pd.DataFrame) -> pd.DataFrame:
    validate_columns(df)
    work = fill_required_fields(df)
    work["_RankNumeric"] = pd.to_numeric(work["Rank"], errors="coerce")
    work["_RaceIdentity"] = work.apply(
        lambda row: race_identity(row.get("race_id", ""), row.get("race_id_name", "")),
        axis=1,
    )
    work["_RaceGroupKey"] = work["_RaceIdentity"].astype("string").fillna("")
    blank_race_mask = work["_RaceGroupKey"].str.strip() == ""
    work.loc[blank_race_mask, "_RaceGroupKey"] = (
        "__row__" + work.index[blank_race_mask].astype(str)
    )

    for col in OUTPUT_COLUMNS:
        work[col] = pd.NA
    work["Final_Rank"] = work["Rank"]

    ordered = work

    env = trueskill.TrueSkill(
        mu=25.0,
        sigma=8.333,
        beta=KEIRIN_BETA,
        tau=KEIRIN_TAU,
    )
    rider_ratings: dict[str, trueskill.Rating] = {}

    for _, race_df in ordered.groupby("_RaceGroupKey", sort=False, dropna=False):
        race_sorted = race_df.sort_values(by="_RankNumeric", kind="mergesort")

        athletes = race_sorted["Athlete"].astype(str).tolist()
        idxs = race_sorted.index.tolist()
        ranks_series = race_sorted["Rank"]
        dnf_dns_mask = ranks_series.apply(is_excluded_rank)

        initial_ratings = [
            rider_ratings.get(athlete, env.create_rating()) for athlete in athletes
        ]
        initial_map = {idx: rating for idx, rating in zip(idxs, initial_ratings)}
        athlete_map = {idx: athlete for idx, athlete in zip(idxs, athletes)}
        eligible_idxs = race_sorted[~dnf_dns_mask].index.tolist()
        eligible_athletes = [athlete_map[idx] for idx in eligible_idxs]
        eligible_initial_ratings = [initial_map[idx] for idx in eligible_idxs]
        eligible_ranks = race_sorted.loc[eligible_idxs, "Rank"]

        strength = pd.NA
        if eligible_initial_ratings:
            strength = float(
                sum(r.mu for r in eligible_initial_ratings) / len(eligible_initial_ratings)
            )

        for idx, rating in zip(idxs, initial_ratings):
            work.at[idx, "Initial_Mu"] = float(rating.mu)
            work.at[idx, "Initial_Sigma"] = float(rating.sigma)
            work.at[idx, "Initial_CSE"] = float(rating.mu - 3 * rating.sigma)
            work.at[idx, "Strength_of_Field"] = strength

        # DNF/DNS/REL rows are excluded from this race's rating update.
        # Their ratings remain unchanged in rider_ratings until a future eligible result.

        # TrueSkill requires at least two competitors in a match.
        if len(eligible_initial_ratings) < 2:
            for idx in eligible_idxs:
                rating = initial_map[idx]
                athlete = athlete_map[idx]
                rider_ratings[athlete] = rating
                work.at[idx, "Final_Mu"] = float(rating.mu)
                work.at[idx, "Final_Sigma"] = float(rating.sigma)
                work.at[idx, "Final_CSE"] = float(rating.mu - 3 * rating.sigma)
            continue

        groups = [[rating] for rating in eligible_initial_ratings]
        ranks = rank_values_for_trueskill(eligible_ranks)
        updated_groups = env.rate(groups, ranks=ranks)

        for idx, athlete, updated_group in zip(
            eligible_idxs, eligible_athletes, updated_groups
        ):
            rating = updated_group[0]
            rider_ratings[athlete] = rating

            work.at[idx, "Final_Mu"] = float(rating.mu)
            work.at[idx, "Final_Sigma"] = float(rating.sigma)
            work.at[idx, "Final_CSE"] = float(rating.mu - 3 * rating.sigma)

    work = work.drop(columns=["_RankNumeric", "_RaceIdentity", "_RaceGroupKey"])
    return work


def normalize_race_id(value: object) -> str:
    if pd.isna(value):
        return ""

    if isinstance(value, (int, float)):
        num = float(value)
        if num.is_integer():
            return str(int(num))
        return format(num, ".15g")

    text = str(value).strip()
    if not text:
        return ""

    # Normalize text-like numeric IDs (e.g. "123" and "123.0").
    try:
        num = float(text)
        if num.is_integer():
            return str(int(num))
        return format(num, ".15g")
    except ValueError:
        return text


def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip().lower()


def race_identity(race_id: object, race_id_name: object) -> str:
    race_id_norm = normalize_race_id(race_id)
    if race_id_norm:
        return race_id_norm
    return normalize_text(race_id_name)


def row_key(race_id: object, race_id_name: object, athlete: object, rank: object) -> str:
    race_norm = race_identity(race_id, race_id_name)
    if not race_norm:
        return ""
    athlete_norm = normalize_text(athlete)
    rank_norm = normalize_text(rank)
    return f"{race_norm}|{athlete_norm}|{rank_norm}"


def existing_row_keys_from_sheet(ws) -> set[str]:
    if ws.max_row < 2:
        return set()

    header_cells = [cell.value for cell in ws[1]]
    header_lookup = {
        str(value).strip().lower(): idx + 1
        for idx, value in enumerate(header_cells)
        if value is not None
    }

    if "athlete" not in header_lookup or "rank" not in header_lookup:
        return set()
    if "race_id" not in header_lookup and "race_id_name" not in header_lookup:
        return set()

    race_idx = header_lookup.get("race_id", 0) - 1
    race_name_idx = header_lookup.get("race_id_name", 0) - 1
    athlete_idx = header_lookup["athlete"] - 1
    rank_idx = header_lookup["rank"] - 1

    keys: set[str] = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        race_id_value = row[race_idx].value if race_idx >= 0 else ""
        race_id_name_value = row[race_name_idx].value if race_name_idx >= 0 else ""
        race_norm = race_identity(race_id_value, race_id_name_value)
        key = row_key(
            race_id_value,
            race_id_name_value,
            row[athlete_idx].value,
            row[rank_idx].value,
        )
        if race_norm:
            keys.add(key)

    return keys


def append_to_sheet(
    workbook_path: Path,
    output_df: pd.DataFrame,
    output_sheet: str,
    color_scale_exclude: tuple[str, ...] = (),
) -> tuple[int, int]:
    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    wb = load_workbook(workbook_path, keep_vba=keep_vba)

    if output_sheet not in wb.sheetnames:
        ws = wb.create_sheet(output_sheet)
        ws.append(output_df.columns.tolist())
    else:
        ws = wb[output_sheet]
        if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
            ws.delete_rows(1)
            ws.append(output_df.columns.tolist())

    existing_row_keys = existing_row_keys_from_sheet(ws)

    output_df = output_df.copy()
    output_df["_race_identity"] = output_df.apply(
        lambda row: race_identity(row.get("race_id", ""), row.get("race_id_name", "")),
        axis=1,
    )
    output_df["_row_key"] = output_df.apply(
        lambda row: row_key(row.get("race_id", ""), row.get("race_id_name", ""), row["Athlete"], row["Rank"]),
        axis=1,
    )

    # Keep all rows except exact duplicates where a reliable race key exists.
    dedupe_mask = (output_df["_row_key"] != "") & output_df["_row_key"].isin(existing_row_keys)
    filtered_df = output_df[~dedupe_mask]

    filtered_df = filtered_df.drop(columns=["_race_identity", "_row_key"])

    # Write rows using worksheet header order so values land in the intended columns.
    header_cells = [cell.value for cell in ws[1]]
    header_names = ["" if value is None else str(value).strip() for value in header_cells]
    header_lookup = {name.lower(): idx for idx, name in enumerate(header_names) if name}

    # Rename filtered_df columns to match worksheet casing so reindex matches exactly.
    df_rename_map: dict[str, str] = {}
    for col in list(filtered_df.columns):
        sheet_idx = header_lookup.get(col.lower())
        if sheet_idx is not None:
            sheet_name = header_names[sheet_idx]
            if sheet_name and sheet_name != col:
                df_rename_map[col] = sheet_name
    if df_rename_map:
        filtered_df = filtered_df.rename(columns=df_rename_map)

    # If sheet has an Age-like header (e.g. 'Age (yrs)') but no plain 'Age', map our Age column to it.
    if "Age" in filtered_df.columns and "age" not in header_lookup:
        age_alias = next(
            (name for name in header_names if name and name.lower().startswith("age")),
            None,
        )
        if age_alias:
            filtered_df = filtered_df.rename(columns={"Age": age_alias})

    for col in filtered_df.columns:
        if col.lower() not in header_lookup:
            ws.cell(row=1, column=len(header_names) + 1, value=col)
            header_names.append(col)
            header_lookup[col.lower()] = len(header_names) - 1

    normalized_to_actual = {name.lower(): name for name in header_names if name}
    ordered_columns = [
        normalized_to_actual[name.lower()]
        for name in header_names
        if name and name.lower() in normalized_to_actual
    ]

    aligned_df = filtered_df.reindex(columns=ordered_columns)

    round_col_idx = header_lookup.get("round", -1)
    heat_col_idx = header_lookup.get("heat", -1)

    skipped_rows = len(output_df) - len(filtered_df)
    rows_written = 0
    for row in dataframe_to_rows(aligned_df, index=False, header=False):
        next_row = ws.max_row + 1
        ws.append([None if pd.isna(v) else v for v in row])
        if round_col_idx >= 0 and round_col_idx < len(row):
            heat_value = row[heat_col_idx] if heat_col_idx >= 0 and heat_col_idx < len(row) else 1
            row_fill = keirin_row_fill(row[round_col_idx], heat_value)
            if row_fill is not None:
                for col_idx in range(1, len(header_names) + 1):
                    ws.cell(row=next_row, column=col_idx).fill = row_fill
        rows_written += 1

    apply_metric_color_scales(ws, header_lookup, exclude_columns=color_scale_exclude)

    wb.save(workbook_path)
    return rows_written, skipped_rows


METRIC_COLOR_SCALE_COLUMNS = (
    "Initial_Mu",
    "Initial_Sigma",
    "Initial_CSE",
    "Final_Mu",
    "Final_Sigma",
    "Final_CSE",
    "Strength_of_Field",
)


def apply_metric_color_scales(
    ws,
    header_lookup: dict[str, int],
    exclude_columns: tuple[str, ...] = (),
) -> None:
    if ws.max_row < 2:
        return

    last_row = ws.max_row
    excluded = {name.lower() for name in exclude_columns}

    for column_name in METRIC_COLOR_SCALE_COLUMNS:
        if column_name.lower() in excluded:
            continue
        idx = header_lookup.get(column_name.lower())
        if idx is None:
            continue

        column_letter = get_column_letter(idx + 1)
        cell_range = f"{column_letter}2:{column_letter}{last_row}"

        if column_name.endswith("Sigma"):
            # Lower sigma is better (more certainty) -> green low, red high.
            rule = ColorScaleRule(
                start_type="min", start_color="FF63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFFFEB84",
                end_type="max", end_color="FFF8696B",
            )
        else:
            # Higher is better -> red low, green high.
            rule = ColorScaleRule(
                start_type="min", start_color="FFF8696B",
                mid_type="percentile", mid_value=50, mid_color="FFFFEB84",
                end_type="max", end_color="FF63BE7B",
            )

        existing_rules = ws.conditional_formatting._cf_rules  # type: ignore[attr-defined]
        for existing_range in list(existing_rules.keys()):
            if str(existing_range) == cell_range:
                del existing_rules[existing_range]

        ws.conditional_formatting.add(cell_range, rule)


def process_workbook(
    workbook: Path,
    input_sheet: str,
    output_sheet: str,
    color_scale_exclude: tuple[str, ...] = (),
) -> tuple[int, int]:
    input_df = pd.read_excel(workbook, sheet_name=input_sheet)
    output_df = compute_keirin_trueskill(input_df)

    # Force Age to come straight from the Keirin tab (lookup by athlete + race).
    age_col = next(
        (c for c in input_df.columns if str(c).strip().lower() == "age"),
        None,
    )
    if age_col is not None:
        src = input_df.copy()
        src["_race_key"] = src.apply(
            lambda r: race_identity(r.get("race_id", ""), r.get("race_id_name", "")),
            axis=1,
        )
        # If the race_identity is blank in source, fall back to a row-built name.
        blank_mask = src["_race_key"].str.strip() == ""
        if blank_mask.any():
            src.loc[blank_mask, "_race_key"] = src.loc[blank_mask].apply(
                _build_race_id_name_from_row, axis=1
            ).str.lower().str.strip()
        src["_athlete_key"] = src.get("Athlete", "").astype(str).str.strip().str.lower()
        src["_age_value"] = src[age_col]
        age_lookup = (
            src.dropna(subset=["_age_value"])
            .drop_duplicates(subset=["_race_key", "_athlete_key"], keep="last")
            .set_index(["_race_key", "_athlete_key"])["_age_value"]
        )

        def _lookup_age(row: pd.Series) -> object:
            race_key = race_identity(row.get("race_id", ""), row.get("race_id_name", ""))
            if not race_key:
                race_key = _build_race_id_name_from_row(row).lower().strip()
            athlete_key = str(row.get("Athlete", "")).strip().lower()
            try:
                return age_lookup.loc[(race_key, athlete_key)]
            except KeyError:
                return pd.NA

        output_df["Age"] = output_df.apply(_lookup_age, axis=1)

    return append_to_sheet(
        workbook, output_df, output_sheet, color_scale_exclude=color_scale_exclude
    )


def resolve_workbook_path(workbook: Path) -> Path:
    if workbook.is_absolute() or workbook.exists():
        return workbook

    resolved = CNZPD_DIR / workbook
    return resolved


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Calculate Keirin TrueSkill and append only new race_id rows into "
            "Keirin_Trueskill for both mens and womens workbooks."
        )
    )
    parser.add_argument(
        "--mens-workbook",
        type=Path,
        default=Path("pages/MensRaceResults.xlsm"),
        help="Path to mens workbook (default: pages/MensRaceResults.xlsm)",
    )
    parser.add_argument(
        "--womens-workbook",
        type=Path,
        default=Path("pages/WomensRaceResults.xlsm"),
        help="Path to womens workbook (default: pages/WomensRaceResults.xlsm)",
    )
    parser.add_argument(
        "--input-sheet",
        default="Keirin",
        help="Input sheet name (default: Keirin)",
    )
    parser.add_argument(
        "--output-sheet",
        default="Keirin_Trueskill",
        help="Output sheet name (default: Keirin_Trueskill)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_configs = [
        ("Mens", resolve_workbook_path(args.mens_workbook)),
        ("Womens", resolve_workbook_path(args.womens_workbook)),
    ]

    results: list[str] = []
    for label, workbook in workbook_configs:
        exclude = ("Strength_of_Field",) if label == "Mens" else ()
        written, skipped = process_workbook(
            workbook,
            args.input_sheet,
            args.output_sheet,
            color_scale_exclude=exclude,
        )
        results.append(f"{label}: appended {written}, skipped {skipped}.")

    print("Done. " + " ".join(results))


if __name__ == "__main__":
    main()