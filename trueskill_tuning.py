"""Chronological TrueSkill tuning for sprint and keirin results.

This module keeps the event order intact, supports separate sprint and
keirin evaluation, and exposes a grid-search pipeline that scores only after
a configurable burn-in period.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Literal, Sequence

import pandas as pd
import trueskill
from openpyxl import load_workbook


EventType = Literal["sprint", "keirin"]

DEFAULT_MU = 25.0
DEFAULT_SIGMA = DEFAULT_MU / 3.0
BASE_DIR = Path(__file__).resolve().parent

WORKBOOK_TUNING_TARGETS = (
    {
        "workbook_path": BASE_DIR / "pages/MensRaceResults.xlsm",
        "event_type": "sprint",
        "source_sheet": "Sprint",
        "target_sheet": "Sprint_Trueskill",
        "beta_values": [x / 2 for x in range(1, 21)],
    },
    {
        "workbook_path": BASE_DIR / "pages/MensRaceResults.xlsm",
        "event_type": "keirin",
        "source_sheet": "Keirin",
        "target_sheet": "Keirin_Trueskill",
    },
    {
        "workbook_path": BASE_DIR / "pages/WomensRaceResults.xlsm",
        "event_type": "sprint",
        "source_sheet": "Sprint",
        "target_sheet": "Sprint_Trueskill",
    },
    {
        "workbook_path": BASE_DIR / "pages/WomensRaceResults.xlsm",
        "event_type": "keirin",
        "source_sheet": "Keirin",
        "target_sheet": "Keirin_Trueskill",
    },
)


@dataclass(frozen=True)
class EvaluationSummary:
    """Aggregate metrics for one chronological pass."""

    pairwise_accuracy: float
    winner_accuracy: float
    scored_events: int
    total_events: int


def build_environment(
    beta: float,
    tau: float,
    draw_probability: float = 0.0,
    mu: float = DEFAULT_MU,
    sigma: float = DEFAULT_SIGMA,
) -> trueskill.TrueSkill:
    """Create a TrueSkill environment for the supplied hyperparameters."""

    return trueskill.TrueSkill(
        mu=mu,
        sigma=sigma,
        beta=beta,
        tau=tau,
        draw_probability=draw_probability,
    )


def chronological_split(
    events: Sequence[Sequence[str]],
    train_fraction: float = 0.67,
) -> tuple[list[Sequence[str]], list[Sequence[str]]]:
    """Split a chronologically ordered event list without shuffling."""

    if not 0.0 < train_fraction < 1.0:
        raise ValueError("train_fraction must be between 0 and 1")

    split_index = max(1, min(len(events) - 1, int(round(len(events) * train_fraction))))
    return list(events[:split_index]), list(events[split_index:])


def recommend_burn_in(event_count: int, event_type: EventType) -> int:
    """Return a practical starting burn-in for the given event count.

    Sprint needs a slightly longer warm-up because each match only updates two
    riders. Keirin learns faster because each race updates six riders in one go.
    """

    if event_count <= 0:
        return 0

    if event_type == "sprint":
        return min(max(50, int(round(event_count * 0.05))), max(0, event_count // 3))

    return min(max(20, int(round(event_count * 0.03))), max(0, event_count // 3))


def default_parameter_grid(event_type: EventType) -> tuple[list[float], list[float]]:
    """Return a sensible starting grid for beta and tau."""

    default_beta = DEFAULT_MU / 6.0

    if event_type == "sprint":
        beta_multipliers = [0.5, 0.75, 1.0, 1.25, 1.5, 1.75, 2.0]
        tau_values = [0.0, 0.05, 0.1, 0.2, 0.35, 0.5]
    else:
        beta_multipliers = [0.5, 1.0, 1.25, 1.5, 1.75, 2.0, 2.5, 3.0]
        tau_values = [0.0, 0.05, 0.1, 0.2, 0.35, 0.5, 0.75]

    beta_values = [round(default_beta * multiplier, 4) for multiplier in beta_multipliers]
    return beta_values, tau_values


def pairwise_accuracy(predicted_order: Sequence[str], actual_order: Sequence[str]) -> float:
    """Measure how often the predicted ranking preserves the observed order."""

    if len(actual_order) < 2:
        return 0.0

    actual_positions = {participant_id: index for index, participant_id in enumerate(actual_order)}
    correct_pairs = 0
    total_pairs = 0

    for left_index, left_participant in enumerate(predicted_order):
        for right_participant in predicted_order[left_index + 1 :]:
            total_pairs += 1
            if actual_positions[left_participant] < actual_positions[right_participant]:
                correct_pairs += 1

    return correct_pairs / total_pairs if total_pairs else 0.0


def _float_range(start: float, stop: float, step: float, precision: int = 6) -> list[float]:
    if step <= 0:
        raise ValueError("step must be > 0")
    values: list[float] = []
    value = start
    epsilon = step / 10.0
    while value <= stop + epsilon:
        values.append(round(value, precision))
        value += step
    return sorted(set(values))


def _sorted_unique(values: Iterable[float]) -> list[float]:
    return sorted({float(value) for value in values})


def _search_grid(
    events: Sequence[Sequence[str]],
    event_type: EventType,
    beta_values: Sequence[float],
    tau_values: Sequence[float],
    train_fraction: float,
    burn_in: int | None,
    draw_probability: float,
    stage_label: str,
) -> tuple[pd.DataFrame, pd.Series, int, int]:
    if not events:
        raise ValueError("events must not be empty")

    train_events, test_events = chronological_split(events, train_fraction=train_fraction)
    train_burn_in = burn_in if burn_in is not None else recommend_burn_in(len(train_events), event_type)
    test_burn_in = min(train_burn_in, max(0, len(test_events) - 1)) if test_events else 0

    records: list[dict[str, float | str]] = []
    for beta in beta_values:
        for tau in tau_values:
            environment = build_environment(beta=beta, tau=tau, draw_probability=draw_probability)

            train_summary, train_ratings = evaluate_sequence(
                train_events,
                event_type,
                environment,
                burn_in=train_burn_in,
            )
            test_summary, _ = evaluate_sequence(
                test_events,
                event_type,
                environment,
                burn_in=test_burn_in,
                start_ratings=train_ratings,
            )

            records.append(
                {
                    "stage": stage_label,
                    "beta": float(beta),
                    "tau": float(tau),
                    "train_pairwise_accuracy": train_summary.pairwise_accuracy,
                    "train_winner_accuracy": train_summary.winner_accuracy,
                    "test_pairwise_accuracy": test_summary.pairwise_accuracy,
                    "test_winner_accuracy": test_summary.winner_accuracy,
                    "train_scored_events": float(train_summary.scored_events),
                    "test_scored_events": float(test_summary.scored_events),
                    "train_burn_in": float(train_burn_in),
                    "test_burn_in": float(test_burn_in),
                }
            )

    results = pd.DataFrame.from_records(records)
    results = results.sort_values(
        by=["test_pairwise_accuracy", "test_winner_accuracy", "train_pairwise_accuracy", "beta", "tau"],
        ascending=[False, False, False, True, True],
        kind="mergesort",
    ).reset_index(drop=True)

    best_row = results.iloc[0]
    return results, best_row, train_burn_in, test_burn_in


def stage_results_side_by_side(results: pd.DataFrame, top_n: int = 10) -> pd.DataFrame:
    """Return top Stage 1 and Stage 2 rows side by side for quick comparison."""

    stage1_top = results[results["stage"] == "stage1"].head(top_n).reset_index(drop=True)
    stage2_top = results[results["stage"] == "stage2"].head(top_n).reset_index(drop=True)

    stage1_top = stage1_top.add_prefix("stage1_")
    stage2_top = stage2_top.add_prefix("stage2_")
    return pd.concat([stage1_top, stage2_top], axis=1)


def _best_row_by_test_accuracy(results: pd.DataFrame) -> pd.Series:
    ordered = results.sort_values(
        by=["test_pairwise_accuracy", "test_winner_accuracy", "train_pairwise_accuracy", "beta", "tau"],
        ascending=[False, False, False, True, True],
        kind="mergesort",
    )
    return ordered.iloc[0]


def _ensure_rating(
    ratings: dict[str, trueskill.Rating],
    participant_id: str,
    environment: trueskill.TrueSkill,
) -> trueskill.Rating:
    if participant_id not in ratings:
        ratings[participant_id] = environment.create_rating()
    return ratings[participant_id]


def _predict_order(
    ratings: dict[str, trueskill.Rating],
    event: Sequence[str],
) -> list[str]:
    return sorted(event, key=lambda participant_id: ratings[participant_id].mu, reverse=True)


def _update_ratings(
    ratings: dict[str, trueskill.Rating],
    event: Sequence[str],
    environment: trueskill.TrueSkill,
    event_type: EventType,
) -> None:
    if event_type == "sprint":
        if len(event) != 2:
            raise ValueError("Sprint events must contain exactly two riders")

        winner_id, loser_id = event
        updated_winner, updated_loser = environment.rate_1vs1(
            ratings[winner_id],
            ratings[loser_id],
        )
        ratings[winner_id] = updated_winner
        ratings[loser_id] = updated_loser
        return

    ranked_groups = [[ratings[participant_id]] for participant_id in event]
    updated_groups = environment.rate(
        ranked_groups,
        ranks=list(range(len(event))),
    )
    for participant_id, updated_group in zip(event, updated_groups):
        ratings[participant_id] = updated_group[0]


def _clean_participant_id(value: object) -> str | None:
    if pd.isna(value):
        return None
    cleaned = str(value).strip()
    return cleaned or None


def _choose_group_columns(frame: pd.DataFrame, candidates: Sequence[str]) -> list[str]:
    return [column for column in candidates if column in frame.columns]


def _build_race_id(frame: pd.DataFrame, event_type: EventType) -> pd.Series:
    if event_type == "sprint":
        key_columns = ["Location", "Year", "Event", "Stage", "Heat"]
    else:
        key_columns = ["Location", "Year", "Event", "Round", "Heat"]

    missing_columns = [column for column in key_columns if column not in frame.columns]
    if missing_columns:
        raise ValueError(f"Missing columns needed to build race id: {', '.join(missing_columns)}")

    normalized = frame[key_columns].astype(str).fillna("")
    return normalized.agg("_".join, axis=1)


def _group_event_rows(frame: pd.DataFrame, event_type: EventType) -> list[tuple[object, pd.DataFrame]]:
    race_id_column = _build_race_id(frame, event_type)
    frame = frame.copy()
    frame["derived_race_id"] = race_id_column
    return list(frame.groupby("derived_race_id", dropna=False, sort=False))


def load_events_from_workbook(
    workbook_path: str | Path,
    event_type: EventType,
    source_sheet: str | None = None,
) -> list[tuple[str, ...]]:
    """Load chronologically ordered event sequences from a race-results workbook."""

    workbook_path = Path(workbook_path)
    if source_sheet is None:
        source_sheet = "Sprint" if event_type == "sprint" else "Keirin"

    frame = pd.read_excel(workbook_path, engine="openpyxl", sheet_name=source_sheet)
    frame = frame.replace({"": pd.NA, " ": pd.NA})

    if event_type == "sprint":
        rank_column = "Final_Rank"
    else:
        rank_column = "Rank"

    if rank_column not in frame.columns:
        raise ValueError(f"{source_sheet} must contain a {rank_column} column")

    if "Athlete" not in frame.columns:
        raise ValueError(f"{source_sheet} must contain an Athlete column")

    events: list[tuple[str, ...]] = []
    grouped = frame.dropna(subset=[rank_column, "Athlete"]).copy()
    grouped["derived_race_id"] = _build_race_id(grouped, event_type)

    grouped[rank_column] = pd.to_numeric(grouped[rank_column], errors="coerce")
    grouped = grouped.dropna(subset=[rank_column])

    grouped = grouped.sort_values(["derived_race_id", rank_column])

    for _, race_rows in _group_event_rows(grouped, event_type):
        ordered_rows = race_rows.sort_values(rank_column, ascending=True)
        ordered_participants = [
            _clean_participant_id(participant)
            for participant in ordered_rows["Athlete"].tolist()
        ]
        ordered_participants = [participant for participant in ordered_participants if participant is not None]

        if event_type == "sprint":
            if len(ordered_participants) >= 2:
                events.append((ordered_participants[0], ordered_participants[1]))
            continue

        if len(ordered_participants) >= 2:
            events.append(tuple(ordered_participants))

    return events


def write_grid_search_results_to_workbook(
    workbook_path: str | Path,
    results: pd.DataFrame,
    sheet_name: str = "TrueSkill_Grid_Search",
) -> None:
    """Write a grid-search result table into a new sheet in an .xlsm workbook."""

    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path, keep_vba=True)

    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    worksheet = workbook.create_sheet(title=sheet_name)

    for row_index, row_values in enumerate([list(results.columns)] + results.astype(object).where(pd.notna(results), None).values.tolist(), start=1):
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    workbook.save(workbook_path)


def write_model_summary_to_sheet(
    workbook_path: str | Path,
    sheet_name: str,
    summary: dict[str, object],
) -> None:
    """Write a compact model summary block to the right of an existing sheet."""

    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path, keep_vba=True)

    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"{sheet_name} does not exist in {workbook_path.name}")

    worksheet = workbook[sheet_name]
    start_col = worksheet.max_column + 2

    for row_index, (header, value) in enumerate(summary.items(), start=1):
        worksheet.cell(row=row_index, column=start_col, value=header)
        worksheet.cell(row=row_index, column=start_col + 1, value=value)

    workbook.save(workbook_path)


def tune_workbook_event(
    workbook_path: str | Path,
    event_type: EventType,
    source_sheet: str | None = None,
    target_sheet: str | None = None,
    beta_values: Iterable[float] | None = None,
    tau_values: Iterable[float] | None = None,
    train_fraction: float = 0.67,
    burn_in: int | None = None,
    draw_probability: float = 0.0,
    write_summary: bool = True,
    write_grid_results: bool = True,
    use_two_stage: bool | None = None,
) -> tuple[pd.DataFrame, pd.Series, int, int]:
    """Tune one workbook event type and optionally write the findings back out."""

    workbook_path = Path(workbook_path)
    if use_two_stage is None:
        use_two_stage = event_type == "sprint"

    events = load_events_from_workbook(
        workbook_path=workbook_path,
        event_type=event_type,
        source_sheet=source_sheet,
    )

    results, best_row, train_burn_in, test_burn_in = grid_search_true_skill(
        events=events,
        event_type=event_type,
        beta_values=beta_values,
        tau_values=tau_values,
        train_fraction=train_fraction,
        burn_in=burn_in,
        draw_probability=draw_probability,
        two_stage=use_two_stage,
    )

    stage1_best = _best_row_by_test_accuracy(results[results["stage"] == "stage1"])
    stage2_rows = results[results["stage"] == "stage2"]
    if len(stage2_rows) > 0:
        stage2_best = _best_row_by_test_accuracy(stage2_rows)
        stage2_skipped = False
    else:
        stage2_best = stage1_best
        stage2_skipped = True

    if stage2_skipped or float(stage1_best["test_pairwise_accuracy"]) >= float(stage2_best["test_pairwise_accuracy"]):
        final_best = stage1_best
    else:
        final_best = stage2_best

    if write_grid_results:
        write_grid_search_results_to_workbook(
            workbook_path=workbook_path,
            results=results,
            sheet_name="TrueSkill_Grid_Search",
        )

    if write_summary:
        if target_sheet is None:
            target_sheet = "Sprint_Trueskill" if event_type == "sprint" else "Keirin_Trueskill"

        summary = {
            "Event_Type": event_type,
            "Stage1_Best_Beta": float(stage1_best["beta"]),
            "Stage1_Best_Tau": float(stage1_best["tau"]),
            "Stage1_Train_Acc": float(stage1_best["train_pairwise_accuracy"]),
            "Stage1_Test_Acc": float(stage1_best["test_pairwise_accuracy"]),
            "Stage2_Best_Beta": float(stage2_best["beta"]),
            "Stage2_Best_Tau": float(stage2_best["tau"]),
            "Stage2_Train_Acc": float(stage2_best["train_pairwise_accuracy"]),
            "Stage2_Test_Acc": float(stage2_best["test_pairwise_accuracy"]),
            "Stage2_Skipped": str(stage2_skipped),
            "Final_Beta": float(final_best["beta"]),
            "Final_Tau": float(final_best["tau"]),
            "Final_Test_Acc": float(final_best["test_pairwise_accuracy"]),
            "Train_Burn_In": int(train_burn_in),
            "Test_Burn_In": int(test_burn_in),
            "Scored_Train_Events": int(final_best["train_scored_events"]),
            "Scored_Test_Events": int(final_best["test_scored_events"]),
        }
        write_model_summary_to_sheet(
            workbook_path=workbook_path,
            sheet_name=target_sheet,
            summary=summary,
        )

    return results, best_row, train_burn_in, test_burn_in


def run_all_workbook_tuning(
    train_fraction: float = 0.67,
    draw_probability: float = 0.0,
    write_grid_results: bool = True,
    write_summary: bool = True,
    print_stage_tables: bool = True,
) -> pd.DataFrame:
    """Tune every workbook/event combination and write results back out."""

    records: list[dict[str, object]] = []

    for target in WORKBOOK_TUNING_TARGETS:
        results, best_row, train_burn_in, test_burn_in = tune_workbook_event(
            workbook_path=target["workbook_path"],
            event_type=target["event_type"],
            source_sheet=target["source_sheet"],
            target_sheet=target["target_sheet"],
            beta_values=target.get("beta_values"),
            tau_values=target.get("tau_values"),
            train_fraction=train_fraction,
            draw_probability=draw_probability,
            write_grid_results=write_grid_results,
            write_summary=write_summary,
            use_two_stage=(target["event_type"] == "sprint"),
        )
        if print_stage_tables:
            label = f"{target['workbook_path'].name} - {target['event_type']}"
            print(f"\n=== {label} ===")
            print(stage_results_side_by_side(results, top_n=8).to_string(index=False))

        stage1_best = _best_row_by_test_accuracy(results[results["stage"] == "stage1"])
        stage2_rows = results[results["stage"] == "stage2"]
        if len(stage2_rows) > 0:
            stage2_best = _best_row_by_test_accuracy(stage2_rows)
            stage2_skipped = False
        else:
            stage2_best = stage1_best
            stage2_skipped = True

        if stage2_skipped or float(stage1_best["test_pairwise_accuracy"]) >= float(stage2_best["test_pairwise_accuracy"]):
            final_best = stage1_best
        else:
            final_best = stage2_best
        records.append(
            {
                "workbook": target["workbook_path"].as_posix(),
                "event_type": target["event_type"],
                "stage1_beta": float(stage1_best["beta"]),
                "stage1_tau": float(stage1_best["tau"]),
                "stage1_train_acc": float(stage1_best["train_pairwise_accuracy"]),
                "stage1_test_acc": float(stage1_best["test_pairwise_accuracy"]),
                "stage2_beta": float(stage2_best["beta"]),
                "stage2_tau": float(stage2_best["tau"]),
                "stage2_train_acc": float(stage2_best["train_pairwise_accuracy"]),
                "stage2_test_acc": float(stage2_best["test_pairwise_accuracy"]),
                "stage2_skipped": stage2_skipped,
                "final_beta": float(final_best["beta"]),
                "final_tau": float(final_best["tau"]),
                "final_test_acc": float(final_best["test_pairwise_accuracy"]),
                "train_burn_in": int(train_burn_in),
                "test_burn_in": int(test_burn_in),
                "scored_train_events": int(final_best["train_scored_events"]),
                "scored_test_events": int(final_best["test_scored_events"]),
            }
        )

    return pd.DataFrame.from_records(records)


def evaluate_sequence(
    events: Sequence[Sequence[str]],
    event_type: EventType,
    environment: trueskill.TrueSkill,
    burn_in: int = 0,
    start_ratings: dict[str, trueskill.Rating] | None = None,
) -> tuple[EvaluationSummary, dict[str, trueskill.Rating]]:
    """Evaluate one chronologically ordered sequence and update ratings in place."""

    ratings = dict(start_ratings or {})
    pairwise_score = 0.0
    winner_score = 0.0
    scored_events = 0

    for event_index, event in enumerate(events):
        actual_order = list(event)

        for participant_id in actual_order:
            _ensure_rating(ratings, participant_id, environment)

        predicted_order = _predict_order(ratings, actual_order)

        if event_index >= burn_in:
            pairwise_score += pairwise_accuracy(predicted_order, actual_order)
            winner_score += 1.0 if predicted_order[0] == actual_order[0] else 0.0
            scored_events += 1

        _update_ratings(ratings, actual_order, environment, event_type)

    summary = EvaluationSummary(
        pairwise_accuracy=pairwise_score / scored_events if scored_events else 0.0,
        winner_accuracy=winner_score / scored_events if scored_events else 0.0,
        scored_events=scored_events,
        total_events=len(events),
    )
    return summary, ratings


def grid_search_true_skill(
    events: Sequence[Sequence[str]],
    event_type: EventType,
    beta_values: Iterable[float] | None = None,
    tau_values: Iterable[float] | None = None,
    train_fraction: float = 0.67,
    burn_in: int | None = None,
    draw_probability: float = 0.0,
    two_stage: bool = True,
    stage1_beta_values: Iterable[float] | None = None,
    stage1_tau_values: Iterable[float] | None = None,
    stage2_beta_step: float = 0.1,
    stage2_tau_step: float = 0.01,
    stage2_beta_window: float = 1.0,
    stage2_tau_window: float = 0.1,
) -> tuple[pd.DataFrame, pd.Series, int, int]:
    """Run a chronological grid search and return the full result table.

    With two_stage=True, Stage 1 performs a coarse wide search and Stage 2
    performs a fine search centered around the Stage 1 best point.

    Stage-level best rows are selected by test pairwise accuracy, and final
    recommended parameters are whichever stage has the higher test accuracy.
    """

    if not events:
        raise ValueError("events must not be empty")

    if not two_stage:
        if beta_values is None or tau_values is None:
            default_beta_values, default_tau_values = default_parameter_grid(event_type)
            beta_values = default_beta_values if beta_values is None else beta_values
            tau_values = default_tau_values if tau_values is None else tau_values

        return _search_grid(
            events=events,
            event_type=event_type,
            beta_values=_sorted_unique(beta_values),
            tau_values=_sorted_unique(tau_values),
            train_fraction=train_fraction,
            burn_in=burn_in,
            draw_probability=draw_probability,
            stage_label="stage1",
        )

    if stage1_beta_values is None:
        if beta_values is not None:
            stage1_beta_values = beta_values
        else:
            stage1_beta_values = _float_range(1.0, 10.0, 0.5, precision=4)
    if stage1_tau_values is None:
        if tau_values is not None:
            stage1_tau_values = tau_values
        else:
            stage1_tau_values = _float_range(0.0, 1.0, 0.05, precision=4)

    stage1_results, stage1_best, train_burn_in, test_burn_in = _search_grid(
        events=events,
        event_type=event_type,
        beta_values=_sorted_unique(stage1_beta_values),
        tau_values=_sorted_unique(stage1_tau_values),
        train_fraction=train_fraction,
        burn_in=burn_in,
        draw_probability=draw_probability,
        stage_label="stage1",
    )

    stage2_beta_min = max(0.01, float(stage1_best["beta"]) - stage2_beta_window)
    stage2_beta_max = float(stage1_best["beta"]) + stage2_beta_window
    stage2_tau_min = max(0.0, float(stage1_best["tau"]) - stage2_tau_window)
    stage2_tau_max = max(stage2_tau_min, float(stage1_best["tau"]) + stage2_tau_window)

    stage2_beta_values = _float_range(stage2_beta_min, stage2_beta_max, stage2_beta_step, precision=6)
    stage2_tau_values = _float_range(stage2_tau_min, stage2_tau_max, stage2_tau_step, precision=6)

    stage2_results, stage2_best, _, _ = _search_grid(
        events=events,
        event_type=event_type,
        beta_values=stage2_beta_values,
        tau_values=stage2_tau_values,
        train_fraction=train_fraction,
        burn_in=burn_in,
        draw_probability=draw_probability,
        stage_label="stage2",
    )

    all_results = pd.concat([stage1_results, stage2_results], ignore_index=True)
    if float(stage1_best["test_pairwise_accuracy"]) >= float(stage2_best["test_pairwise_accuracy"]):
        best_row = stage1_best
    else:
        best_row = stage2_best

    return all_results, best_row, train_burn_in, test_burn_in


def tune_sprint_and_keirin(
    sprint_events: Sequence[Sequence[str]],
    keirin_events: Sequence[Sequence[str]],
    sprint_beta_values: Iterable[float] | None = None,
    sprint_tau_values: Iterable[float] | None = None,
    keirin_beta_values: Iterable[float] | None = None,
    keirin_tau_values: Iterable[float] | None = None,
    train_fraction: float = 0.67,
    sprint_burn_in: int | None = None,
    keirin_burn_in: int | None = None,
    draw_probability: float = 0.0,
) -> dict[str, object]:
    """Tune sprint and keirin separately and return both result tables."""

    sprint_results, sprint_best, sprint_train_burn_in, sprint_test_burn_in = grid_search_true_skill(
        sprint_events,
        event_type="sprint",
        beta_values=sprint_beta_values,
        tau_values=sprint_tau_values,
        train_fraction=train_fraction,
        burn_in=sprint_burn_in,
        draw_probability=draw_probability,
    )
    keirin_results, keirin_best, keirin_train_burn_in, keirin_test_burn_in = grid_search_true_skill(
        keirin_events,
        event_type="keirin",
        beta_values=keirin_beta_values,
        tau_values=keirin_tau_values,
        train_fraction=train_fraction,
        burn_in=keirin_burn_in,
        draw_probability=draw_probability,
    )

    return {
        "sprint_results": sprint_results,
        "sprint_best": sprint_best,
        "sprint_train_burn_in": sprint_train_burn_in,
        "sprint_test_burn_in": sprint_test_burn_in,
        "keirin_results": keirin_results,
        "keirin_best": keirin_best,
        "keirin_train_burn_in": keirin_train_burn_in,
        "keirin_test_burn_in": keirin_test_burn_in,
    }


def plot_grid_search_results(
    results: pd.DataFrame,
    metric_prefix: str = "pairwise",
    title_prefix: str = "TrueSkill",
) -> dict[str, object]:
    """Visualise training and test accuracy over the parameter grid.

    Returns pandas Styler objects that render as heatmaps in notebooks or can
    be converted to HTML for reporting.
    """

    train_metric = f"train_{metric_prefix}_accuracy"
    test_metric = f"test_{metric_prefix}_accuracy"

    if train_metric not in results.columns or test_metric not in results.columns:
        raise ValueError(f"Results must contain {train_metric} and {test_metric}")

    train_pivot = results.pivot(index="tau", columns="beta", values=train_metric).sort_index()
    test_pivot = results.pivot(index="tau", columns="beta", values=test_metric).sort_index()

    train_heatmap = train_pivot.style.background_gradient(cmap="YlGnBu", axis=None, vmin=0.0, vmax=1.0).format(
        "{:.3f}"
    )
    test_heatmap = test_pivot.style.background_gradient(cmap="YlGnBu", axis=None, vmin=0.0, vmax=1.0).format(
        "{:.3f}"
    )
    train_heatmap.set_caption(f"{title_prefix} Train {metric_prefix.replace('_', ' ').title()} Accuracy")
    test_heatmap.set_caption(f"{title_prefix} Test {metric_prefix.replace('_', ' ').title()} Accuracy")

    return {"train": train_heatmap, "test": test_heatmap}


__all__ = [
    "EvaluationSummary",
    "build_environment",
    "chronological_split",
    "default_parameter_grid",
    "evaluate_sequence",
    "grid_search_true_skill",
    "load_events_from_workbook",
    "pairwise_accuracy",
    "plot_grid_search_results",
    "recommend_burn_in",
    "run_all_workbook_tuning",
    "stage_results_side_by_side",
    "tune_workbook_event",
    "write_model_summary_to_sheet",
    "write_grid_search_results_to_workbook",
    "tune_sprint_and_keirin",
]


if __name__ == "__main__":
    summary_frame = run_all_workbook_tuning()
    print(summary_frame.to_string(index=False))