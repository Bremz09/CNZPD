from __future__ import annotations

import argparse
import shutil
import tempfile
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule


SCRIPT_DIR = Path(__file__).resolve().parent
SOURCE_DIR = SCRIPT_DIR / "CNZPD" / "pages"
TARGET_DIR = SOURCE_DIR / "WR_progressions"

SOURCE_FILES = {
    "Mens": SOURCE_DIR / "MensRaceResults.xlsm",
    "Womens": SOURCE_DIR / "WomensRaceResults.xlsm",
}
TARGET_FILES = {
    "Mens": TARGET_DIR / "Mens_Progression.xlsx",
    "Womens": TARGET_DIR / "Womens_Progression.xlsx",
}

# These source sheets have row-oriented result data that can be projected into
# the existing progression sheets without inventing aggregate values.
ROW_MAPPINGS = {
    "Madison": "Medals_Madison",
    "OM-Scratch": "Medals_Om_Scratch",
    "OM-Tempo": "Medals_Om_Tempo",
    "OM-Elimination": "Medals_Om_Elim",
    "OM-Points": "Medals_Om_Points",
}

RESULT_LIMITS = {
    "Madison": 3,
    "OM-Scratch": 1,
    "OM-Elimination": 1,
    "OM-Tempo": 3,
    "OM-Points": 3,
}

ALLOWED_EVENTS = {"NC", "WCH", "OLY"}
ROW_EVENT_FILTERS = {"Madison": {"NC", "WCH", "OLY"}, "OM-Scratch": {"NC", "WCH", "OLY"}, "OM-Tempo": {"NC", "WCH", "OLY"}, "OM-Elimination": {"NC", "WCH", "OLY"}, "OM-Points": {"NC", "WCH", "OLY"}}
AGGREGATE_EVENT_FILTERS = {"F200": {"NC", "WCH", "OLY"}, "TS": {"NC", "WCH", "OLY"}, "TP": {"NC", "WCH", "OLY"}, "IP": {"NC", "WCH", "OLY"}}

AGGREGATE_MAPPINGS = {
    "Sprint Qual": ("Medals_F200", "F200"),
    "Team Sprint": ("Medals_TS", "TS"),
    "Team Pursuit": ("Medals_TP", "TP"),
    "Individual Pursuit": ("Medals_IP", "IP"),
}

HISTORICAL_WOMENS_MAPPINGS = {
    "3k Individual Pursuit": ("Medals_3kIP", "IP"),
}

@dataclass
class SheetResult:
    gender: str
    source_sheet: str
    target_sheet: str
    source_rows: int
    appended_rows: int
    duplicate_rows: int
    skipped_rows: int
    missing_columns: list[str]


def clean_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    aliases = {"Nat": "Country"}
    df.rename(columns={column: aliases.get(column, column) for column in df.columns}, inplace=True)
    return df


def normalize_key_value(value: Any) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, (datetime, pd.Timestamp)):
        return pd.Timestamp(value).date().isoformat()
    return str(value).strip().casefold()


def row_key(row: pd.Series) -> tuple[str, ...]:
    identity_columns = ["Location", "Year", "Date"]
    identity_columns.append("Event" if "Event" in row.index else "Competition")
    for candidate in ("Name", "Rider1", "Athlete"):
        if candidate in row.index:
            identity_columns.append(candidate)
            break
    if "Rank" in row.index:
        identity_columns.append("Rank")
    return tuple(normalize_key_value(row.get(column)) for column in identity_columns)


def read_source_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    return normalize_columns(pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl"))


def select_progression_rows(source: pd.DataFrame, source_sheet: str) -> pd.DataFrame:
    ranked = source.copy()
    rank = pd.to_numeric(ranked.get("Rank"), errors="coerce")
    selected = rank <= RESULT_LIMITS[source_sheet]
    event_filter = ROW_EVENT_FILTERS.get(source_sheet, ALLOWED_EVENTS)
    selected &= ranked["Event"].astype(str).str.strip().isin(event_filter)

    if source_sheet == "Madison":
        selected &= ranked["Stage"].astype(str).str.strip().eq("F")
        selected &= ranked["Heat"].astype(str).str.strip().eq("F")

    return ranked.loc[selected].copy()


def calculate_omnium_points_fields(source: pd.DataFrame) -> pd.DataFrame:
    calculated = source.copy()
    sprint_columns = [
        column
        for column in calculated.columns
        if str(column).strip().casefold().startswith("sprint ")
    ]
    sprint_values = calculated[sprint_columns].apply(pd.to_numeric, errors="coerce")
    calculated["Sprints Scored"] = sprint_values.gt(0).sum(axis=1)
    calculated["Sprints Won"] = sprint_values.eq(5).sum(axis=1)

    final_sprint = next((column for column in sprint_columns if str(column).strip() == "Sprint 10"), None)
    if final_sprint is not None:
        calculated["Sprints Won"] += sprint_values[final_sprint].eq(10).astype(int)

    final_points = pd.to_numeric(calculated.get("Final"), errors="coerce")
    subtotal = pd.to_numeric(calculated.get("Sub Total"), errors="coerce")
    calculated["Points Total"] = final_points - subtotal
    calculated["P.Laps"] = pd.to_numeric(calculated.get("Lap +"), errors="coerce")
    return calculated


def calculate_omnium_tempo_fields(source: pd.DataFrame) -> pd.DataFrame:
    calculated = source.copy()
    sprint_columns = [
        column
        for column in calculated.columns
        if str(column).strip().isdigit() and 1 <= int(str(column).strip()) <= 36
    ]
    sprint_values = calculated[sprint_columns].apply(pd.to_numeric, errors="coerce")
    calculated["Sprints Won"] = sprint_values.eq(1).sum(axis=1)
    calculated["P.Laps"] = pd.to_numeric(calculated.get("P.Laps"), errors="coerce")
    return calculated


def calculate_madison_fields(source: pd.DataFrame) -> pd.DataFrame:
    calculated = source.copy()
    sprint_columns = [
        column
        for column in calculated.columns
        if str(column).strip().casefold().startswith("sprint ")
    ]
    sprint_values = calculated[sprint_columns].apply(pd.to_numeric, errors="coerce")
    calculated["Sprints Scored"] = sprint_values.gt(0).sum(axis=1)
    calculated["Sprints Won"] = sprint_values.eq(5).sum(axis=1)
    final_sprint = next((column for column in sprint_columns if str(column).strip() == "Sprint 20"), None)
    if final_sprint is not None:
        calculated["Sprints Won"] += sprint_values[final_sprint].eq(10).astype(int)
    calculated["Avg Speed"] = pd.to_numeric(calculated.get("Avg Speed"), errors="coerce")
    return calculated


def select_final_rows(source: pd.DataFrame) -> pd.DataFrame:
    selected = source["Stage"].astype(str).str.strip().eq("F")
    if "Heat" in source.columns:
        final_rows = source.loc[selected]
        if final_rows["Heat"].astype(str).str.strip().eq("F").any():
            selected &= source["Heat"].astype(str).str.strip().eq("F")
    return source.loc[selected].copy()


def seconds(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, time):
        return value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1_000_000
    if isinstance(value, (datetime, pd.Timestamp)):
        value = pd.Timestamp(value)
        return value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1_000_000
    if isinstance(value, (int, float)):
        return float(value)
    parsed = pd.to_timedelta(str(value), errors="coerce")
    return None if pd.isna(parsed) else parsed.total_seconds()


def excel_time(value: Any) -> time | None:
    value_seconds = seconds(value)
    if value_seconds is None:
        return None
    whole_seconds = int(value_seconds)
    return time(
        hour=(whole_seconds // 3600) % 24,
        minute=(whole_seconds % 3600) // 60,
        second=whole_seconds % 60,
        microsecond=round((value_seconds - whole_seconds) * 1_000_000),
    )


def placing_label(placing: int, values: dict[str, Any]) -> str:
    aliases = {
        1: ("1st", "Gold_Time"),
        2: ("2nd", "Silver_Time"),
        3: ("3rd", "Bronze_Time"),
        8: ("8th",),
        16: ("16th",),
    }
    return next((label for label in aliases[placing] if label in values), aliases[placing][0])


def seconds_label(placing: int, values: dict[str, Any]) -> str:
    label = placing_label(placing, values)
    aliases = {
        "Gold_Time": "Gold_Seconds",
        "Silver_Time": "Silver_Seconds",
        "Bronze_Time": "Bronze_Seconds",
    }
    return aliases.get(label, f"{label}_seconds")


def assign_final_group_ranks(group: pd.DataFrame) -> pd.DataFrame:
    ranked = group.copy()
    ranked["_heat_order"] = ranked["Heat"].astype(str).str.strip().map({"G": 0, "B": 1}).fillna(2)
    ranked["_source_rank"] = pd.to_numeric(ranked["Rank"], errors="coerce")
    ranked["_time_seconds"] = ranked["Time"].map(seconds)
    ranked = ranked.sort_values(
        ["_heat_order", "_source_rank", "_time_seconds"],
        na_position="last",
    ).copy()
    ranked["Overall Rank"] = range(1, len(ranked) + 1)
    return ranked


def aggregate_rows(source: pd.DataFrame, target_headers: list[str], event_kind: str) -> pd.DataFrame:
    event_filter = AGGREGATE_EVENT_FILTERS[event_kind]
    source = source[source["Event"].astype(str).str.strip().isin(event_filter)].copy()
    all_source = source
    if event_kind in ("TP", "TS"):
        source["Date"] = pd.to_datetime(source["Date"], errors="coerce")
        source = source[source["Stage"].astype(str).str.strip().eq("F")].copy()
        source = source.loc[
            source.groupby(["Year", "Event", "Location"], dropna=False)["Date"].transform("max").eq(source["Date"])
        ].copy()
    if event_kind == "F200":
        value_column = "200m"
        ranked = source
    elif event_kind in ("TP", "TS"):
        value_column = "Time"
        ranked = select_final_rows(source)
    else:
        value_column = "Time"
        ranked = select_final_rows(source) if "Stage" in source.columns else source

    rank_column = "Rank"
    if rank_column not in ranked.columns and event_kind == "IP" and len(ranked.columns) > 8:
        rank_column = ranked.columns[8]

    group_columns = [column for column in ("Year", "Date", "Event", "Location") if column in source.columns]
    rows: list[dict[str, Any]] = []
    for group_values, group in source.groupby(group_columns, dropna=False, sort=False):
        if not isinstance(group_values, tuple):
            group_values = (group_values,)
        group_info = dict(zip(group_columns, group_values))
        result = {header: None for header in target_headers}
        result["Year"] = group_info.get("Year")
        if "Competition" in result:
            result["Competition"] = group_info.get("Event")
        if "Event" in result:
            result["Event"] = group_info.get("Event")
        if "Location" in result:
            result["Location"] = group_info.get("Location")
        if "Date" in result:
            result["Date"] = group_info.get("Date")
        if "DateSerial" in result and group_info.get("Date") is not None:
            result["DateSerial"] = to_excel(pd.Timestamp(group_info["Date"]).to_pydatetime())

        if event_kind in ("TP", "TS"):
            ranked = assign_final_group_ranks(group)

        for placing in (1, 2, 3, 8, 16):
            if event_kind in ("TP", "TS") and placing in (8, 16):
                continue
            placement_source = source if event_kind == "IP" and placing in (8, 16) else ranked
            if event_kind in ("TP", "TS"):
                match = placement_source.loc[placement_source["Overall Rank"] == placing]
            else:
                match = placement_source.loc[
                    (placement_source["Year"] == group_info.get("Year"))
                    & (placement_source["Event"] == group_info.get("Event"))
                    & (placement_source["Location"] == group_info.get("Location"))
                    & (pd.to_numeric(placement_source[rank_column], errors="coerce") == placing)
                ]
            if match.empty:
                continue
            value = match.iloc[0][value_column]
            label = placing_label(placing, result)
            result[label] = value if event_kind in ("F200", "TS") else excel_time(value)
            seconds_column = seconds_label(placing, result)
            if seconds_column in result:
                value_seconds = seconds(value)
                result[seconds_column] = round(value_seconds, 2) if event_kind in ("TP", "TS") and value_seconds is not None else value_seconds
            if event_kind == "Madison" and placing in (1, 2, 3) and "Avg Speed" in match.columns:
                speed_label = f"Avg Speed" if placing == 1 else f"Avg Speed"
                if speed_label in result:
                    result[speed_label] = match.iloc[0].get("Avg Speed")

        if event_kind in ("TP", "TS"):
            final_date = group_info.get("Date")
            competition_rows = all_source.loc[
                (all_source["Year"] == group_info.get("Year"))
                & (all_source["Event"] == group_info.get("Event"))
                & (all_source["Location"] == group_info.get("Location"))
            ]
            event_day = competition_rows.loc[
                pd.to_datetime(competition_rows["Date"], errors="coerce") == final_date
            ]

            event_times = (competition_rows if event_kind == "TP" else event_day)["Time"].map(seconds).dropna().sort_values()
            if not event_times.empty:
                fastest_seconds = round(float(event_times.iloc[0]), 2)
                if "Fastest" in result and event_kind == "TP":
                    result["Fastest"] = excel_time(fastest_seconds)
                fastest_seconds_column = next(
                    (header for header in result if header.casefold() == "fastest_seconds"),
                    None,
                )
                if fastest_seconds_column is not None and event_kind == "TP":
                    result[fastest_seconds_column] = fastest_seconds

            if event_kind == "TP":
                qualifying_times = competition_rows.loc[
                    competition_rows["Stage"].astype(str).str.strip().eq("Q")
                ]["Time"].map(seconds).dropna().sort_values()
                for index, value in enumerate(qualifying_times.head(3), start=1):
                    label = f"Q{index}"
                    if label in result:
                        result[label] = excel_time(value)
                    seconds_label_name = f"{label}_seconds"
                    if seconds_label_name in result:
                        result[seconds_label_name] = round(float(value), 2)

                qualifying_eighth = competition_rows.loc[
                    competition_rows["Stage"].astype(str).str.strip().eq("Q")
                    & (pd.to_numeric(competition_rows["Rank"], errors="coerce") == 8)
                ].sort_values("Date", ascending=False)
                if not qualifying_eighth.empty:
                    value = seconds(qualifying_eighth.iloc[0]["Time"])
                    if value is not None:
                        if "Q8" in result:
                            result["Q8"] = excel_time(value)
                        if "Q8_seconds" in result:
                            result["Q8_seconds"] = round(value, 2)

            r1 = competition_rows.loc[
                competition_rows["Stage"].astype(str).str.strip().eq("R1")
                & (pd.to_datetime(competition_rows["Date"], errors="coerce") == final_date)
            ]
            r1_times = r1["Time"].map(seconds).dropna()
            if not r1_times.empty:
                value = r1_times.max()
                label = placing_label(8, result)
                result[label] = value if event_kind == "TS" else excel_time(value)
                seconds_column = seconds_label(8, result)
                if seconds_column in result:
                    result[seconds_column] = round(value, 2)

            qualifying = competition_rows.loc[
                competition_rows["Stage"].astype(str).str.strip().eq("Q")
                & (pd.to_numeric(competition_rows["Rank"], errors="coerce") == 16)
            ].sort_values("Date", ascending=False)
            if not qualifying.empty:
                value = seconds(qualifying.iloc[0]["Time"])
                if value is not None:
                    label = placing_label(16, result)
                    result[label] = value if event_kind == "TS" else excel_time(value)
                    seconds_column = seconds_label(16, result)
                    if seconds_column in result:
                        result[seconds_column] = round(value, 2)

        rows.append(result)
    return pd.DataFrame(rows, columns=target_headers)


def format_omnium_points_metrics(worksheet, headers: list[str]) -> None:
    metric_headers = {"Scratch", "Tempo", "Elimination", "Sub Total", "Sprints Scored", "Sprints Won"}
    last_row = worksheet.max_row
    if last_row < 2:
        return

    for column_index, header in enumerate(headers, start=1):
        if header not in metric_headers:
            continue
        column_letter = get_column_letter(column_index)
        cell_range = f"{column_letter}2:{column_letter}{last_row}"
        worksheet.conditional_formatting.add(
            cell_range,
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )


def format_team_pursuit_values(worksheet, headers: list[str]) -> None:
    metadata_headers = {"Location", "Year", "Date", "Event", "DateSerial"}
    time_headers = {
        header
        for header in headers
        if header not in metadata_headers and not header.casefold().endswith("seconds")
    }
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for column_index, cell in enumerate(row):
            header = headers[column_index] if column_index < len(headers) else ""
            if header in metadata_headers or cell.value is None:
                continue
            if isinstance(cell.value, (datetime, time, pd.Timestamp)):
                cell.number_format = "m:ss.00"
            elif header in time_headers or isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"


def read_target_headers(path: Path, sheet_name: str) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        headers = [clean_header(cell.value) for cell in worksheet[1]]
        return [header for header in headers if header]
    finally:
        workbook.close()


def row_signature(row: pd.Series, headers: list[str]) -> tuple[str, ...]:
    values: list[str] = []
    for header in headers:
        value = row.get(header)
        if value is None or pd.isna(value):
            values.append("")
        elif header.casefold() == "date":
            values.append(pd.Timestamp(value).date().isoformat())
        elif isinstance(value, (time, datetime, pd.Timestamp)):
            values.append(f"time:{seconds(value):.6f}")
        elif isinstance(value, (int, float)):
            values.append(f"number:{float(value):.6f}")
        else:
            values.append(str(value).strip().casefold())
    return tuple(values)


def read_target_row_signatures(path: Path, sheet_name: str, headers: list[str]) -> set[tuple[str, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        signatures: set[tuple[str, ...]] = set()
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            row = pd.Series(dict(zip(headers, values)))
            if any(normalize_key_value(value) for value in row.values):
                signatures.add(row_signature(row, headers))
        return signatures
    finally:
        workbook.close()


def transform_rows(
    source: pd.DataFrame,
    target_headers: list[str],
    existing_signatures: set[tuple[str, ...]],
) -> tuple[list[list[Any]], int, int, list[str]]:
    def row_cell_value(row: pd.Series, column: str) -> Any:
        value = row.get(column)
        if isinstance(value, pd.Series):
            return value.iloc[0] if not value.empty else None
        return value

    missing_columns = [column for column in target_headers if column not in source.columns]
    usable_headers = [column for column in target_headers if column in source.columns]
    if not usable_headers:
        return [], 0, len(source), missing_columns

    appended: list[list[Any]] = []
    seen_signatures = set(existing_signatures)
    duplicates = 0
    skipped = 0

    for _, source_row in source.iterrows():
        if any(pd.isna(row_cell_value(source_row, column)) for column in usable_headers):
            # Missing values are allowed in fields that are absent from the
            # source schema, but a row with no identity cannot be deduplicated.
            if not any(normalize_key_value(row_cell_value(source_row, column)) for column in ("Date", "Event", "Competition", "Location", "Year")):
                skipped += 1
                continue

        target_row = [row_cell_value(source_row, column) if column in source.columns else None for column in target_headers]
        target_series = pd.Series(dict(zip(target_headers, target_row)))
        signature = row_signature(target_series, target_headers)
        if signature in seen_signatures:
            duplicates += 1
            continue
        seen_signatures.add(signature)
        appended.append(target_row)

    return appended, duplicates, skipped, missing_columns


def process_gender(gender: str, write: bool) -> list[SheetResult]:
    source_path = SOURCE_FILES[gender]
    target_path = TARGET_FILES[gender]
    source_workbook = pd.ExcelFile(source_path, engine="openpyxl")
    results: list[SheetResult] = []
    pending: dict[str, list[list[Any]]] = {}
    tp_qualifying_eighth: dict[str, pd.DataFrame] = {}

    for source_sheet, target_sheet in ROW_MAPPINGS.items():
        if source_sheet not in source_workbook.sheet_names:
            results.append(SheetResult(gender, source_sheet, target_sheet, 0, 0, 0, 0, ["source sheet missing"]))
            continue
        target_headers = read_target_headers(target_path, target_sheet)
        source = read_source_sheet(source_path, source_sheet)
        source_row_count = len(source)
        source = select_progression_rows(source, source_sheet)
        if source_sheet == "OM-Points":
            source = calculate_omnium_points_fields(source)
        elif source_sheet == "OM-Tempo":
            source = calculate_omnium_tempo_fields(source)
        elif source_sheet == "Madison":
            source = calculate_madison_fields(source)
        existing_signatures = read_target_row_signatures(target_path, target_sheet, target_headers)
        rows, duplicates, skipped, missing = transform_rows(source, target_headers, existing_signatures)
        pending[target_sheet] = rows
        results.append(
            SheetResult(
                gender,
                source_sheet,
                target_sheet,
                source_row_count,
                len(rows),
                duplicates,
                skipped + source_row_count - len(source),
                missing,
            )
        )

    for source_sheet, (target_sheet, event_kind) in AGGREGATE_MAPPINGS.items():
        actual_source_sheet = source_sheet
        if actual_source_sheet not in source_workbook.sheet_names:
            results.append(SheetResult(gender, source_sheet, target_sheet, 0, 0, 0, 0, ["source sheet missing"]))
            continue
        target_headers = read_target_headers(target_path, target_sheet)
        source = read_source_sheet(source_path, actual_source_sheet)
        aggregate = aggregate_rows(source, target_headers, event_kind)
        if event_kind == "TP":
            tp_qualifying_eighth[target_sheet] = aggregate
        existing_signatures = read_target_row_signatures(target_path, target_sheet, target_headers)
        rows, duplicates, skipped, missing = transform_rows(aggregate, target_headers, existing_signatures)
        pending[target_sheet] = rows
        results.append(
            SheetResult(
                gender,
                source_sheet,
                target_sheet,
                len(source),
                len(rows),
                duplicates,
                skipped + len(source) - len(aggregate),
                missing,
            )
        )

    if gender == "Womens":
        for source_sheet, (target_sheet, event_kind) in HISTORICAL_WOMENS_MAPPINGS.items():
            if source_sheet not in source_workbook.sheet_names:
                results.append(SheetResult(gender, source_sheet, target_sheet, 0, 0, 0, 0, ["source sheet missing"]))
                continue
            target_headers = read_target_headers(target_path, target_sheet)
            source = read_source_sheet(source_path, source_sheet)
            aggregate = aggregate_rows(source, target_headers, event_kind)
            existing_signatures = read_target_row_signatures(target_path, target_sheet, target_headers)
            rows, duplicates, skipped, missing = transform_rows(aggregate, target_headers, existing_signatures)
            pending[target_sheet] = rows
            results.append(
                SheetResult(
                    gender,
                    source_sheet,
                    target_sheet,
                    len(source),
                    len(rows),
                    duplicates,
                    skipped + len(source) - len(aggregate),
                    missing,
                )
            )

    if write:
        workbook = load_workbook(target_path)
        try:
            for target_sheet, rows in pending.items():
                worksheet = workbook[target_sheet]
                target_headers = read_target_headers(target_path, target_sheet)

                date_column_index = None
                for i, header in enumerate(target_headers):
                    if header.strip().lower() == "date":
                        date_column_index = i
                        break

                if date_column_index is not None and rows:
                    rows_sorted = sorted(
                        rows,
                        key=lambda row: (
                            pd.Timestamp(row[date_column_index]).to_pydatetime()
                            if date_column_index < len(row) and row[date_column_index] is not None
                            else pd.Timestamp.min.to_pydatetime()
                        ),
                        reverse=True,
                    )
                else:
                    rows_sorted = rows

                if rows_sorted:
                    worksheet.insert_rows(2, len(rows_sorted))

                    fill_light = PatternFill(start_color="DAE9F8", end_color="DAE9F8", fill_type="solid")
                    fill_medium = PatternFill(start_color="BED9F4", end_color="BED9F4", fill_type="solid")

                    location_column_index = None
                    for i, header in enumerate(target_headers):
                        if header.strip().lower() == "location":
                            location_column_index = i
                            break

                    current_location = None
                    use_light_color = True

                    for idx, row in enumerate(rows_sorted, start=2):
                        if location_column_index is not None and location_column_index < len(row):
                            row_location = row[location_column_index]
                            if row_location != current_location:
                                current_location = row_location
                                use_light_color = not use_light_color

                        current_fill = fill_light if use_light_color else fill_medium

                        for col_idx, value in enumerate(row, start=1):
                            cell = worksheet.cell(row=idx, column=col_idx, value=value)
                            cell.fill = current_fill
                            if date_column_index is not None and col_idx == date_column_index + 1:
                                if value is not None and not pd.isna(value):
                                    cell.number_format = "mm/dd/yyyy"

                if target_sheet in tp_qualifying_eighth:
                    qualifying_rows = tp_qualifying_eighth[target_sheet]
                    updates_by_key = {
                        row_key(row): row
                        for _, row in qualifying_rows.iterrows()
                    }
                    q8_column_index = target_headers.index("Q8") + 1 if "Q8" in target_headers else None
                    q8_seconds_column_index = (
                        target_headers.index("Q8_seconds") + 1
                        if "Q8_seconds" in target_headers
                        else None
                    )
                    for cells in worksheet.iter_rows(min_row=2):
                        existing_row = pd.Series(
                            dict(zip(target_headers, (cell.value for cell in cells)))
                        )
                        qualifying_row = updates_by_key.get(row_key(existing_row))
                        if qualifying_row is None:
                            continue
                        if q8_column_index is not None:
                            worksheet.cell(cells[0].row, q8_column_index).value = qualifying_row.get("Q8")
                        if q8_seconds_column_index is not None:
                            worksheet.cell(cells[0].row, q8_seconds_column_index).value = qualifying_row.get("Q8_seconds")

                if target_sheet == "Medals_Om_Points":
                    format_omnium_points_metrics(worksheet, target_headers)
                elif target_sheet == "Medals_TP":
                    format_team_pursuit_values(worksheet, target_headers)

            with tempfile.NamedTemporaryFile(delete=False, suffix=target_path.suffix, dir=target_path.parent) as temporary:
                temporary_path = Path(temporary.name)
            workbook.save(temporary_path)
            shutil.copy2(temporary_path, target_path)
            temporary_path.unlink(missing_ok=True)
        finally:
            workbook.close()

    source_workbook.close()
    return results


def main() -> None:
    parser = argparse.ArgumentParser(description="Append senior race results to progression workbooks.")
    parser.add_argument("--write", dest="write", action="store_true", default=True, help="Write changes (the default).")
    parser.add_argument("--dry-run", dest="write", action="store_false", help="Preview changes without modifying workbooks.")
    args = parser.parse_args()

    mode = "WRITE" if args.write else "DRY-RUN"
    print(f"Mode: {mode}")
    for gender in SOURCE_FILES:
        results = process_gender(gender, write=args.write)
        for result in results:
            missing = ", ".join(result.missing_columns) if result.missing_columns else "none"
            print(
                f"{result.gender}: {result.source_sheet} -> {result.target_sheet} | "
                f"source={result.source_rows} appended={result.appended_rows} "
                f"duplicates={result.duplicate_rows} skipped={result.skipped_rows} "
                f"notes={missing}"
            )


if __name__ == "__main__":
    main()
