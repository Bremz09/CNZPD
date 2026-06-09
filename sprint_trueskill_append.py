from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Dict, Iterable, List, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from trueskill import Rating, TrueSkill


STAGE_ORDER = ["R32", "Rep32", "R16", "Rep16", "R8", "Rep8", "QF", "5to8", "SF", "F"]
MIN_INITIAL_CSE = 0.001

PAIRWISE_FINAL_STAGES = {"F"}

# Explicit TrueSkill environment for reproducibility.
TS_MU = 25.0
TS_SIGMA = 8.333333333
TS_DRAW_PROBABILITY = 0.0
MENS_SPRINT_BETA = 2.0
MENS_SPRINT_TAU = 0.20
WOMENS_SPRINT_BETA = 6.0
WOMENS_SPRINT_TAU = 0.75


SPRINT_STAGE_FILL_MAP = {
    "R32": PatternFill(fill_type="solid", start_color="FFE2EFDA", end_color="FFE2EFDA"),
    "REP32": PatternFill(fill_type="solid", start_color="FFE2EFDA", end_color="FFE2EFDA"),
    "R16": PatternFill(fill_type="solid", start_color="FFE2EFDA", end_color="FFE2EFDA"),
    "REP16": PatternFill(fill_type="solid", start_color="FFE2EFDA", end_color="FFE2EFDA"),
    "R8": PatternFill(fill_type="solid", start_color="FFFCE4D6", end_color="FFFCE4D6"),
    "QF": PatternFill(fill_type="solid", start_color="FFD0CECE", end_color="FFD0CECE"),
    "SF": PatternFill(fill_type="solid", start_color="FFFFF2CC", end_color="FFFFF2CC"),
    "F": PatternFill(fill_type="solid", start_color="FFD9E1F2", end_color="FFD9E1F2"),
    "FINAL": PatternFill(fill_type="solid", start_color="FFD9E1F2", end_color="FFD9E1F2"),
}


def normalize_stage(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip().upper()
    return re.sub(r"\s+", "", text)


METRIC_COLOR_SCALE_COLUMNS = (
    "Initial_Mu",
    "Initial_Sigma",
    "Initial_CSE",
    "Final_Mu",
    "Final_Sigma",
    "Final_CSE",
    "Strength_of_Field",
)


def apply_metric_color_scales(ws, header_to_col: dict[str, int]) -> None:
    if ws.max_row < 2:
        return

    last_row = ws.max_row
    normalized_lookup = {str(name).strip().lower(): idx for name, idx in header_to_col.items()}

    for column_name in METRIC_COLOR_SCALE_COLUMNS:
        idx = normalized_lookup.get(column_name.lower())
        if idx is None:
            continue

        column_letter = get_column_letter(idx)
        cell_range = f"{column_letter}2:{column_letter}{last_row}"

        if column_name.endswith("Sigma"):
            rule = ColorScaleRule(
                start_type="min", start_color="FF63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFFFEB84",
                end_type="max", end_color="FFF8696B",
            )
        else:
            rule = ColorScaleRule(
                start_type="min", start_color="FFF8696B",
                mid_type="percentile", mid_value=50, mid_color="FFFFEB84",
                end_type="max", end_color="FF63BE7B",
            )

        existing_rules = ws.conditional_formatting._cf_rules  # type: ignore[attr-defined]
        for existing_range in list(existing_rules.keys()):
            if str(existing_range) == cell_range:
                del existing_rules[existing_range]

        ws.conditional_formatting.add(cell_range, rule)


def build_trueskill_env(beta: float, tau: float) -> TrueSkill:
    return TrueSkill(
        mu=TS_MU,
        sigma=TS_SIGMA,
        beta=beta,
        tau=tau,
        draw_probability=TS_DRAW_PROBABILITY,
    )


def _string_part(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return text


def _int_string_part(value: object) -> str:
    if pd.isna(value):
        return ""
    try:
        return str(int(float(value)))
    except Exception:  # noqa: BLE001
        return _string_part(value)


@dataclass
class WorkbookConfig:
    path: Path
    sprint_sheet: str
    trueskill_sheet: str


def normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


def find_sheet_name(candidates: Iterable[str], target: str) -> str | None:
    target_norm = normalize_name(target)
    for item in candidates:
        if normalize_name(item) == target_norm:
            return item
    return None


def find_required_column(df: pd.DataFrame, options: Sequence[str]) -> str:
    normalized = {normalize_name(col): col for col in df.columns}
    for option in options:
        match = normalized.get(normalize_name(option))
        if match:
            return match
    raise KeyError(f"Could not find any of columns: {options}")


def read_workbook_config(path: Path) -> WorkbookConfig:
    wb = load_workbook(path, read_only=True, keep_vba=True)
    try:
        sprint_sheet = find_sheet_name(wb.sheetnames, "Sprint")
        ts_sheet = find_sheet_name(wb.sheetnames, "Sprint_Trueskill")
        if not sprint_sheet:
            raise ValueError(f"Sprint sheet not found in {path}")
        if not ts_sheet:
            raise ValueError(f"Sprint_Trueskill sheet not found in {path}")
        return WorkbookConfig(path=path, sprint_sheet=sprint_sheet, trueskill_sheet=ts_sheet)
    finally:
        wb.close()


def read_sheet_as_dataframe(path: Path, sheet_name: str) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, keep_vba=True, data_only=True)
    try:
        ws = wb[sheet_name]
        values = list(ws.iter_rows(values_only=True))
        if not values:
            return pd.DataFrame()
        headers = [str(h) if h is not None else "" for h in values[0]]
        rows = values[1:]
        return pd.DataFrame(rows, columns=headers)
    finally:
        wb.close()


def prepare_sprint_rows(df_raw: pd.DataFrame) -> pd.DataFrame:
    df = df_raw.copy()

    col_rank_r1 = find_required_column(df, ["Rank R1", "Rank_R1"])
    col_rank_r2 = find_required_column(df, ["Rank R2", "Rank_R2"])
    col_rank_r3 = find_required_column(df, ["Rank R3", "Rank_R3"])
    col_time_r1 = find_required_column(df, ["200m R1", "200m_R1"])
    col_time_r2 = find_required_column(df, ["200m R2", "200m_R2"])
    col_time_r3 = find_required_column(df, ["200m R3", "200m_R3"])
    col_speed_r1 = find_required_column(df, ["Avg Speed R1", "Avg_Speed_R1"])
    col_speed_r2 = find_required_column(df, ["Avg Speed R2", "Avg_Speed_R2"])
    col_speed_r3 = find_required_column(df, ["Avg Speed R3", "Avg_Speed_R3"])

    df.insert(4 if len(df.columns) >= 4 else len(df.columns), "Round", 1)
    df.insert(15 if len(df.columns) >= 15 else len(df.columns), "Avg Speed", df[col_speed_r1])

    df2 = df.loc[pd.to_numeric(df[col_rank_r2], errors="coerce") > 0].copy()
    df3 = df.loc[pd.to_numeric(df[col_rank_r3], errors="coerce") > 0].copy()

    if not df2.empty:
        df2["Round"] = 2
        df2[col_rank_r1] = df2[col_rank_r2]
        df2[col_time_r1] = df2[col_time_r2]
        df2["Avg Speed"] = df2[col_speed_r2]

    if not df3.empty:
        df3["Round"] = 3
        df3[col_rank_r1] = df3[col_rank_r3]
        df3[col_time_r1] = df3[col_time_r3]
        df3["Avg Speed"] = df3[col_speed_r3]

    df_all = pd.concat([df, df2, df3], ignore_index=True)
    df_all.insert(11 if len(df_all.columns) >= 11 else len(df_all.columns), "Rank", df_all[col_rank_r1])
    df_all.insert(12 if len(df_all.columns) >= 12 else len(df_all.columns), "Time", df_all[col_time_r1])

    drop_cols = [col_rank_r1, col_time_r1, col_speed_r1, col_rank_r2, col_time_r2, col_speed_r2, col_rank_r3, col_time_r3, col_speed_r3]
    df_all = df_all.drop(columns=[c for c in drop_cols if c in df_all.columns])

    # Provide aliases so output columns named '200m' or 'Gap' get populated.
    if "Time" in df_all.columns and "200m" not in df_all.columns:
        df_all["200m"] = df_all["Time"]

    for metric_col in [
        "Initial_CSE",
        "Initial_Mu",
        "Initial_Sigma",
        "Final_CSE",
        "Final_Mu",
        "Final_Sigma",
        "Strength_of_Field",
    ]:
        if metric_col not in df_all.columns:
            df_all[metric_col] = pd.NA

    if "Stage" in df_all.columns:
        df_all["Stage"] = pd.Categorical(df_all["Stage"], STAGE_ORDER, ordered=True)

    sort_cols = [c for c in ["Date", "Stage", "Round", "Heat", "Rank"] if c in df_all.columns]
    if sort_cols:
        df_all = df_all.sort_values(sort_cols)

    if "Rank" in df_all.columns:
        df_all = df_all[pd.to_numeric(df_all["Rank"], errors="coerce") > 0].copy()

    # Build a deterministic race context label requested for append/dedupe.
    race_name_parts = [
        df_all["Location"].map(_string_part) if "Location" in df_all.columns else "",
        df_all["Year"].map(_int_string_part) if "Year" in df_all.columns else "",
        df_all["Event"].map(_string_part) if "Event" in df_all.columns else "",
        df_all["Stage"].map(_string_part) if "Stage" in df_all.columns else "",
        df_all["Heat"].map(_int_string_part) if "Heat" in df_all.columns else "",
        df_all["Round"].map(_int_string_part) if "Round" in df_all.columns else "",
    ]

    race_id_name = race_name_parts[0].astype("string")
    for part in race_name_parts[1:]:
        race_id_name = race_id_name + " " + part.astype("string")

    df_all["race_id_name"] = race_id_name.str.replace(r"\s+", " ", regex=True).str.strip()

    # Compute Gap (time behind winner) within each race so output sheets have it filled.
    if "Time" in df_all.columns:
        time_numeric = pd.to_numeric(df_all["Time"], errors="coerce")
        winner_time = time_numeric.groupby(df_all["race_id_name"]).transform("min")
        gap = (time_numeric - winner_time).round(4)
        df_all["Gap"] = gap

    return df_all.reset_index(drop=True)


def fill_required_fields(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()

    # Ensure race_id_name follows: location year event stage heat round.
    race_name_parts = [
        work["Location"].map(_string_part) if "Location" in work.columns else "",
        work["Year"].map(_int_string_part) if "Year" in work.columns else "",
        work["Event"].map(_string_part) if "Event" in work.columns else "",
        work["Stage"].map(_string_part) if "Stage" in work.columns else "",
        work["Heat"].map(_int_string_part) if "Heat" in work.columns else "",
        work["Round"].map(_int_string_part) if "Round" in work.columns else "",
    ]

    race_id_name = race_name_parts[0].astype("string")
    for part in race_name_parts[1:]:
        race_id_name = race_id_name + " " + part.astype("string")
    race_id_name = race_id_name.str.replace(r"\s+", " ", regex=True).str.strip()

    if "race_id" in work.columns:
        race_id = work["race_id"].astype("string").fillna("").str.strip()
        race_id_name = race_id_name.where(race_id_name != "", race_id)

    work["race_id_name"] = race_id_name

    return work


def build_group_columns(df: pd.DataFrame) -> List[str]:
    priority = ["Date", "Location", "Year", "Event", "Stage", "Round"]
    group_cols = [c for c in priority if c in df.columns]
    if not group_cols:
        fallback = [c for c in ["Date", "Location", "Event", "Stage", "Round"] if c in df.columns]
        group_cols = fallback if fallback else ["Round"]
    return group_cols


def calculate_trueskill(df: pd.DataFrame, env: TrueSkill) -> pd.DataFrame:
    athlete_col = find_required_column(df, ["Athlete"])
    rank_col = find_required_column(df, ["Rank"])

    ratings: Dict[str, Rating] = {}

    def _update_group(race_df: pd.DataFrame) -> None:
        nonlocal df, ratings
        race_df = race_df.copy()
        race_df[rank_col] = pd.to_numeric(race_df[rank_col], errors="coerce")
        race_df = race_df.dropna(subset=[athlete_col, rank_col])
        race_df = race_df[race_df[rank_col] > 0]
        race_df = race_df.sort_values(rank_col)
        if race_df.empty:
            return

        participants = race_df[athlete_col].astype(str).tolist()
        pre_ratings = [ratings.get(name, env.create_rating()) for name in participants]
        ranks = [int(x) - 1 for x in race_df[rank_col].tolist()]

        for row_index, name, r_pre in zip(race_df.index.tolist(), participants, pre_ratings):
            df.at[row_index, "Initial_Mu"] = float(r_pre.mu)
            df.at[row_index, "Initial_Sigma"] = float(r_pre.sigma)
            initial_cse = float(r_pre.mu - 3.0 * r_pre.sigma)
            df.at[row_index, "Initial_CSE"] = initial_cse if initial_cse > MIN_INITIAL_CSE else MIN_INITIAL_CSE

        # TrueSkill requires at least two competitors in a race group.
        # For singleton groups, carry rating through unchanged.
        if len(participants) < 2:
            r_only = pre_ratings[0]
            row_index = race_df.index.tolist()[0]
            df.at[row_index, "Final_Mu"] = float(r_only.mu)
            df.at[row_index, "Final_Sigma"] = float(r_only.sigma)
            df.at[row_index, "Final_CSE"] = float(r_only.mu - 3.0 * r_only.sigma)
            df.at[row_index, "Strength_of_Field"] = float(r_only.mu)
            ratings[participants[0]] = r_only
            return

        # Sort by rank
        race_df = race_df.sort_values(rank_col)

        # Only use winner (rank 1) and last place.
        # Choose a distinct loser athlete when possible so loser writeback is not overwritten.
        winner_name = str(race_df.iloc[0][athlete_col])
        winner_row_index = race_df.index[0]

        loser_row_index = race_df.index[-1]
        loser_name = str(race_df.iloc[-1][athlete_col])
        if loser_name == winner_name and len(race_df) > 1:
            for idx in race_df.index[::-1]:
                candidate_name = str(race_df.at[idx, athlete_col])
                if candidate_name != winner_name:
                    loser_row_index = idx
                    loser_name = candidate_name
                    break

        winner_pre = ratings.get(winner_name, env.create_rating())
        loser_pre = ratings.get(loser_name, env.create_rating())

        rating_groups = [(winner_pre,), (loser_pre,)]
        ranks = [0, 1]

        rated = env.rate(rating_groups, ranks=ranks)

        # Write back only winner and loser
        ratings[winner_name] = rated[0][0]
        ratings[loser_name] = rated[1][0]

        strength_of_field = float(sum(r.mu for r in pre_ratings) / len(pre_ratings))

        winner_post = rated[0][0]
        loser_post = rated[1][0]

        df.loc[winner_row_index, "Final_Mu"] = float(winner_post.mu)
        df.loc[winner_row_index, "Final_Sigma"] = float(winner_post.sigma)
        df.loc[winner_row_index, "Final_CSE"] = float(winner_post.mu - 3.0 * winner_post.sigma)
        df.loc[winner_row_index, "Strength_of_Field"] = strength_of_field

        df.loc[loser_row_index, "Final_Mu"] = float(loser_post.mu)
        df.loc[loser_row_index, "Final_Sigma"] = float(loser_post.sigma)
        df.loc[loser_row_index, "Final_CSE"] = float(loser_post.mu - 3.0 * loser_post.sigma)
        df.loc[loser_row_index, "Strength_of_Field"] = strength_of_field

    base_cols = [c for c in ["Date", "Location", "Year", "Event", "Stage", "Round"] if c in df.columns]
    if not base_cols:
        base_cols = [c for c in ["Date", "Event", "Stage", "Round"] if c in df.columns]
    if not base_cols:
        base_cols = ["Round"]

    for _, stage_df in df.groupby(base_cols, sort=False, dropna=False):
        # Sprint races are commonly represented as separate heats within a stage.
        # Score each heat independently so both rider outputs are produced per race.
        if "Heat" in stage_df.columns:
            for _, heat_df in stage_df.groupby(["Heat"], sort=False, dropna=False):
                _update_group(heat_df)
        else:
            _update_group(stage_df)

    return df


def _make_dedupe_key(df: pd.DataFrame) -> pd.Series:
    # Duplicate rule requested: same race id + same rider => skip.
    # Same race id with a different rider must still append.
    def norm_series(col_name: str) -> pd.Series:
        if col_name not in df.columns:
            return pd.Series([""] * len(df), index=df.index, dtype="string")
        return df[col_name].astype("string").fillna("").str.strip().str.lower()

    # Preferred race identity: race_id_name, then race_id, then explicit context columns.
    context_cols = [c for c in ["Location", "Year", "Event", "Stage", "Heat", "Round"] if c in df.columns]
    if context_cols:
        race_key = norm_series(context_cols[0])
        for col in context_cols[1:]:
            race_key = race_key + "|" + norm_series(col)
    else:
        race_key = pd.Series([""] * len(df), index=df.index, dtype="string")

    race_id = norm_series("race_id")
    race_id_name = norm_series("race_id_name")

    fallback_parts = []
    for candidate in ["Date", "Location", "Year", "Event", "Stage", "Heat", "Round"]:
        if candidate in df.columns:
            fallback_parts.append(norm_series(candidate))

    if fallback_parts:
        fallback_race = fallback_parts[0]
        for part in fallback_parts[1:]:
            fallback_race = fallback_race + "|" + part
    else:
        fallback_race = pd.Series([""] * len(df), index=df.index, dtype="string")

    # Coalesce per row: race_id_name -> race_id -> context columns -> fallback components.
    context_race = race_key
    race_key = race_id_name.where(race_id_name != "", race_id)
    race_key = race_key.where(race_key != "", context_race)
    race_key = race_key.where(race_key != "", fallback_race)

    rider_col = "Athlete" if "Athlete" in df.columns else None
    if rider_col is None:
        for candidate in ["Rider", "Name"]:
            if candidate in df.columns:
                rider_col = candidate
                break

    if rider_col is None:
        rider_key = pd.Series([""] * len(df), index=df.index, dtype="string")
    else:
        rider_key = df[rider_col].astype("string").fillna("").str.strip().str.lower()

    return race_key + "||" + rider_key


def append_to_sprint_trueskill(workbook_path: Path, results_df: pd.DataFrame, trueskill_sheet: str) -> tuple[int, int]:
    existing_df = read_sheet_as_dataframe(workbook_path, trueskill_sheet)

    existing_keys = set(_make_dedupe_key(existing_df).dropna().tolist()) if not existing_df.empty else set()
    result_keys = _make_dedupe_key(results_df)
    to_append = results_df.loc[~result_keys.isin(existing_keys)].copy()
    total_input_rows = len(results_df)

    # Also guard against duplicate rows within this append batch.
    if not to_append.empty:
        batch_keys = _make_dedupe_key(to_append)
        to_append = to_append.loc[~batch_keys.duplicated()].copy()

    if to_append.empty:
        return 0, total_input_rows

    wb = load_workbook(workbook_path, keep_vba=True)
    try:
        ws = wb[trueskill_sheet]
        headers = [cell.value for cell in ws[1]]
        header_to_col = {str(h): i + 1 for i, h in enumerate(headers) if h is not None}

        missing = [
            c
            for c in [
                "Initial_CSE",
                "Initial_Mu",
                "Initial_Sigma",
                "Final_CSE",
                "Final_Mu",
                "Final_Sigma",
                "Strength_of_Field",
            ]
            if c not in header_to_col
        ]
        for col_name in missing:
            ws.cell(row=1, column=ws.max_column + 1, value=col_name)
            header_to_col[col_name] = ws.max_column

        stage_idx = header_to_col.get("Stage")

        # Build case/space-insensitive lookup so df columns like 'Time' can match
        # sheet headers like '200m', '200m Time', '200m_Time' etc.
        def _norm(name: object) -> str:
            return re.sub(r"[\s_]+", "", str(name)).lower()

        header_norm_to_actual = {_norm(name): name for name in header_to_col}
        header_aliases = {
            "time": ["time", "200m", "200mtime", "200time"],
            "200m": ["200m", "200mtime", "time"],
            "gap": ["gap", "gaptoleader", "gaptowinner"],
            "avgspeed": ["avgspeed", "averagespeed"],
        }

        df_col_to_sheet_col: dict[str, str] = {}
        for df_col in to_append.columns:
            norm = _norm(df_col)
            sheet_col = header_norm_to_actual.get(norm)
            if sheet_col is None:
                for alias in header_aliases.get(norm, []):
                    sheet_col = header_norm_to_actual.get(alias)
                    if sheet_col is not None:
                        break
            if sheet_col is not None:
                df_col_to_sheet_col.setdefault(df_col, sheet_col)

        start_row = ws.max_row + 1
        rows_appended = 0
        for _, row in to_append.iterrows():
            target_row = start_row + rows_appended
            written_sheet_cols: set[str] = set()
            for df_col, sheet_col in df_col_to_sheet_col.items():
                if sheet_col in written_sheet_cols:
                    continue
                value = row[df_col]
                if pd.isna(value):
                    value = None
                ws.cell(row=target_row, column=header_to_col[sheet_col], value=value)
                written_sheet_cols.add(sheet_col)

            if stage_idx is not None and "Stage" in to_append.columns:
                stage_key = normalize_stage(row.get("Stage"))
                fill = SPRINT_STAGE_FILL_MAP.get(stage_key)
                if fill is not None:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=target_row, column=col_idx).fill = fill
            rows_appended += 1

        apply_metric_color_scales(ws, header_to_col)

        wb.save(workbook_path)
        return rows_appended, max(0, total_input_rows - rows_appended)
    finally:
        wb.close()


def process_workbook(path: Path, beta: float, tau: float) -> tuple[int, int]:
    cfg = read_workbook_config(path)
    sprint_df = read_sheet_as_dataframe(path, cfg.sprint_sheet)
    sprint_prepared = prepare_sprint_rows(sprint_df)
    sprint_prepared = fill_required_fields(sprint_prepared)
    sprint_scored = calculate_trueskill(sprint_prepared, build_trueskill_env(beta=beta, tau=tau))
    appended, skipped = append_to_sprint_trueskill(path, sprint_scored, cfg.trueskill_sheet)
    return appended, skipped


def main() -> None:
    base = Path(__file__).resolve().parent
    targets = [
        (base / "MensRaceResults.xlsm", MENS_SPRINT_BETA, MENS_SPRINT_TAU),
        (base / "WomensRaceResults.xlsm", WOMENS_SPRINT_BETA, WOMENS_SPRINT_TAU),
    ]

    for path, beta, tau in targets:
        if not path.exists():
            print(f"Skipped missing workbook: {path}")
            continue

        try:
            added, skipped = process_workbook(path, beta=beta, tau=tau)
            print(
                f"{path.name}: appended {added}, skipped {skipped} row(s) in Sprint_Trueskill"
            )
        except Exception as exc:  # noqa: BLE001
            print(f"{path.name}: failed -> {exc}")


if __name__ == "__main__":
    main()
